"""
Actions module — /api/v1/actions
Track corrective actions, risk treatment plans, CIS remediation tasks.
"""
import io
import os
import uuid
from collections import defaultdict
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import case, delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_session
from app.models.action import Action, ActionAttachment, ActionComment, ActionHistory, ActionLink
from app.models.asset import Asset
from app.models.dictionary import DictionaryEntry
from app.models.org_unit import OrgUnit
from app.models.risk import Risk
from app.schemas.action import (
    ActionAttachmentOut,
    ActionBulkResult,
    ActionBulkUpdate,
    ActionCloseRequest,
    ActionCommentCreate,
    ActionCommentOut,
    ActionCommentUpdate,
    ActionCreate,
    ActionHistoryOut,
    ActionLinkOut,
    ActionMonthlyTrend,
    ActionOut,
    ActionPriorityBreakdown,
    ActionStats,
    ActionStatusBreakdown,
    ActionUpdate,
)

router = APIRouter(prefix="/api/v1/actions", tags=["Dzialania"])

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads", "actions")


# ── helpers ──

async def _de_label(s, entry_id: int | None) -> str | None:
    if entry_id is None:
        return None
    e = await s.get(DictionaryEntry, entry_id)
    return e.label if e else None


async def _entity_name(s, entity_type: str, entity_id: int) -> str | None:
    if entity_type == "risk":
        r = await s.get(Risk, entity_id)
        return r.asset_name if r else None
    if entity_type == "asset":
        a = await s.get(Asset, entity_id)
        return a.name if a else None
    if entity_type == "policy_exception":
        from app.models.policy_exception import PolicyException
        pe = await s.get(PolicyException, entity_id)
        return pe.title if pe else None
    if entity_type == "incident":
        from app.models.incident import Incident
        inc = await s.get(Incident, entity_id)
        return inc.title if inc else None
    if entity_type == "audit":
        from app.models.audit import Audit
        aud = await s.get(Audit, entity_id)
        return aud.title if aud else None
    return f"{entity_type}#{entity_id}"


async def _entity_extra(s, entity_type: str, entity_id: int) -> dict | None:
    """Return extra info about linked entity (e.g. risk score, level, other links)."""
    if entity_type == "risk":
        r = await s.get(Risk, entity_id)
        if not r:
            return None
        other_links_q = select(ActionLink).where(
            ActionLink.entity_type == "risk",
            ActionLink.entity_id == entity_id,
        )
        other_links = (await s.execute(other_links_q)).scalars().all()
        other_action_ids = [lnk.action_id for lnk in other_links]
        status_name = await _de_label(s, r.status_id) if hasattr(r, 'status_id') else None
        return {
            "risk_score": r.risk_score,
            "risk_level": r.risk_level,
            "status_name": status_name,
            "org_unit_name": None,
            "other_action_count": len(other_action_ids),
        }
    return None


async def _action_out(s, action: Action) -> ActionOut:
    org = await s.get(OrgUnit, action.org_unit_id) if action.org_unit_id else None

    # Links
    links_q = select(ActionLink).where(ActionLink.action_id == action.id)
    links = (await s.execute(links_q)).scalars().all()
    link_outs = []
    for lnk in links:
        ename = await _entity_name(s, lnk.entity_type, lnk.entity_id)
        eextra = await _entity_extra(s, lnk.entity_type, lnk.entity_id)
        link_outs.append(ActionLinkOut(
            id=lnk.id, entity_type=lnk.entity_type,
            entity_id=lnk.entity_id, entity_name=ename,
            entity_extra=eextra,
            created_at=lnk.created_at,
        ))

    # History (full, ordered newest first)
    hist_q = (select(ActionHistory)
              .where(ActionHistory.action_id == action.id)
              .order_by(ActionHistory.created_at.desc()))
    history = (await s.execute(hist_q)).scalars().all()

    # Attachments
    att_q = (select(ActionAttachment)
             .where(ActionAttachment.action_id == action.id)
             .order_by(ActionAttachment.created_at.desc()))
    attachments = (await s.execute(att_q)).scalars().all()

    is_overdue = (
        action.due_date is not None
        and action.completed_at is None
        and action.due_date < datetime.utcnow()
    )

    return ActionOut(
        id=action.id,
        title=action.title,
        description=action.description,
        org_unit_id=action.org_unit_id,
        org_unit_name=org.name if org else None,
        owner=action.owner,
        responsible=action.responsible,
        priority_id=action.priority_id,
        priority_name=await _de_label(s, action.priority_id),
        status_id=action.status_id,
        status_name=await _de_label(s, action.status_id),
        source_id=action.source_id,
        source_name=await _de_label(s, action.source_id),
        due_date=action.due_date,
        completed_at=action.completed_at,
        effectiveness_rating=action.effectiveness_rating,
        effectiveness_notes=action.effectiveness_notes,
        implementation_notes=action.implementation_notes,
        is_active=action.is_active,
        is_overdue=is_overdue,
        links=link_outs,
        history=[ActionHistoryOut.model_validate(h) for h in history],
        attachments=[ActionAttachmentOut.model_validate(a) for a in attachments],
        created_at=action.created_at,
        updated_at=action.updated_at,
    )


async def _sync_links(s, action_id: int, links: list[dict]):
    await s.execute(delete(ActionLink).where(ActionLink.action_id == action_id))
    for lnk in links:
        s.add(ActionLink(
            action_id=action_id,
            entity_type=lnk["entity_type"],
            entity_id=lnk["entity_id"],
        ))


async def _track_change(s, action_id: int, field: str, old_val, new_val, change_reason: str | None = None):
    if str(old_val) != str(new_val):
        s.add(ActionHistory(
            action_id=action_id,
            field_name=field,
            old_value=str(old_val) if old_val is not None else None,
            new_value=str(new_val) if new_val is not None else None,
            change_reason=change_reason,
        ))


async def _find_overdue_status_id(s) -> int | None:
    """Find the dictionary entry for 'overdue' action status."""
    from app.models.dictionary import DictionaryType
    q = (
        select(DictionaryEntry.id)
        .select_from(DictionaryEntry)
        .join(DictionaryType, DictionaryEntry.dict_type_id == DictionaryType.id)
        .where(DictionaryType.code == "action_status")
        .where(DictionaryEntry.code == "overdue")
        .limit(1)
    )
    return (await s.execute(q)).scalar()


# ═══════════════════ STATS / KPI ═══════════════════

@router.get("/stats", response_model=ActionStats, summary="Statystyki KPI dzialan")
async def action_stats(s: AsyncSession = Depends(get_session)):
    actions = (await s.execute(select(Action).where(Action.is_active.is_(True)))).scalars().all()
    now = datetime.utcnow()

    total = len(actions)
    completed = [a for a in actions if a.completed_at is not None]
    open_actions = [a for a in actions if a.completed_at is None]
    overdue = [a for a in open_actions if a.due_date is not None and a.due_date < now]

    # Avg completion time
    completion_days = []
    for a in completed:
        if a.completed_at and a.created_at:
            delta = (a.completed_at - a.created_at).days
            completion_days.append(delta)
    avg_completion = sum(completion_days) / len(completion_days) if completion_days else None

    overdue_pct = (len(overdue) / len(open_actions) * 100) if open_actions else 0.0

    # Batch-load all dictionary labels for status/priority
    de_ids = set()
    for a in actions:
        if a.status_id is not None:
            de_ids.add(a.status_id)
        if a.priority_id is not None:
            de_ids.add(a.priority_id)
    de_map: dict[int, str] = {}
    if de_ids:
        rows = (await s.execute(
            select(DictionaryEntry.id, DictionaryEntry.label).where(DictionaryEntry.id.in_(de_ids))
        )).all()
        de_map = {r[0]: r[1] for r in rows}

    # By status
    status_counts: dict[str, int] = defaultdict(int)
    for a in actions:
        sn = de_map.get(a.status_id) if a.status_id else None
        status_counts[sn or "Brak statusu"] += 1
    by_status = [ActionStatusBreakdown(status_name=k, count=v) for k, v in status_counts.items()]

    # By priority
    priority_counts: dict[str, int] = defaultdict(int)
    for a in actions:
        pn = de_map.get(a.priority_id) if a.priority_id else None
        priority_counts[pn or "Brak priorytetu"] += 1
    by_priority = [ActionPriorityBreakdown(priority_name=k, count=v) for k, v in priority_counts.items()]

    # Monthly trend (last 12 months)
    monthly: dict[str, dict[str, int]] = {}
    for i in range(11, -1, -1):
        d = now - timedelta(days=30 * i)
        key = d.strftime("%Y-%m")
        monthly[key] = {"created": 0, "completed": 0}
    for a in actions:
        key = a.created_at.strftime("%Y-%m")
        if key in monthly:
            monthly[key]["created"] += 1
        if a.completed_at:
            ckey = a.completed_at.strftime("%Y-%m")
            if ckey in monthly:
                monthly[ckey]["completed"] += 1
    trend = [ActionMonthlyTrend(month=k, created=v["created"], completed=v["completed"]) for k, v in monthly.items()]

    return ActionStats(
        total=total,
        open=len(open_actions),
        completed=len(completed),
        overdue=len(overdue),
        avg_completion_days=round(avg_completion, 1) if avg_completion is not None else None,
        overdue_pct=round(overdue_pct, 1),
        by_status=by_status,
        by_priority=by_priority,
        monthly_trend=trend,
    )


# ═══════════════════ AUTO-OVERDUE ═══════════════════

@router.post("/auto-overdue", summary="Automatycznie oznacz przeterminowane dzialania")
async def auto_mark_overdue(s: AsyncSession = Depends(get_session)):
    """Find actions past due_date without completed_at and update their status to 'overdue'."""
    try:
        overdue_status_id = await _find_overdue_status_id(s)
        if not overdue_status_id:
            return {"updated_count": 0, "action_ids": []}

        now = datetime.utcnow()
        q = select(Action).where(
            Action.is_active.is_(True),
            Action.due_date < now,
            Action.completed_at.is_(None),
            Action.status_id != overdue_status_id,
        )
        actions = (await s.execute(q)).scalars().all()

        updated_ids = []
        for a in actions:
            await _track_change(s, a.id, "status_id", a.status_id, overdue_status_id,
                                change_reason="Automatyczna zmiana — przekroczono termin realizacji")
            a.status_id = overdue_status_id
            updated_ids.append(a.id)

        if updated_ids:
            await s.commit()

        return {"updated_count": len(updated_ids), "action_ids": updated_ids}
    except Exception:
        await s.rollback()
        return {"updated_count": 0, "action_ids": []}


# ═══════════════════ BULK UPDATE ═══════════════════

@router.post("/bulk", response_model=ActionBulkResult, summary="Masowa zmiana statusu/priorytetu/odpowiedzialnego")
async def bulk_update_actions(body: ActionBulkUpdate, s: AsyncSession = Depends(get_session)):
    q = select(Action).where(Action.id.in_(body.action_ids), Action.is_active.is_(True))
    actions = (await s.execute(q)).scalars().all()

    if not actions:
        raise HTTPException(404, "Nie znaleziono aktywnych dzialan o podanych ID")

    updated_ids = []
    for a in actions:
        changed = False
        if body.status_id is not None and a.status_id != body.status_id:
            await _track_change(s, a.id, "status_id", a.status_id, body.status_id, change_reason=body.change_reason)
            a.status_id = body.status_id
            changed = True
        if body.priority_id is not None and a.priority_id != body.priority_id:
            await _track_change(s, a.id, "priority_id", a.priority_id, body.priority_id, change_reason=body.change_reason)
            a.priority_id = body.priority_id
            changed = True
        if body.responsible is not None and a.responsible != body.responsible:
            await _track_change(s, a.id, "responsible", a.responsible, body.responsible, change_reason=body.change_reason)
            a.responsible = body.responsible
            changed = True
        if changed:
            updated_ids.append(a.id)

    if updated_ids:
        await s.commit()

    return ActionBulkResult(updated_count=len(updated_ids), action_ids=updated_ids)


# ═══════════════════ EXPORT XLSX ═══════════════════

@router.get("/export", summary="Eksport dzialan do Excel (z ryzykami i historia)")
async def export_actions_xlsx(s: AsyncSession = Depends(get_session)):
    """Generate rich XLSX with 3 sheets: Actions, Linked Risks, History."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    actions = (await s.execute(
        select(Action).where(Action.is_active.is_(True)).order_by(
            case((Action.due_date.is_(None), 1), else_=0), Action.due_date.asc()
        )
    )).scalars().all()

    wb = openpyxl.Workbook()

    # ── Sheet 1: Actions ──
    ws1 = wb.active
    ws1.title = "Dzialania"
    headers1 = ["ID", "Tytuł", "Opis", "Jednostka org.", "Właściciel", "Odpowiedzialny",
                 "Priorytet", "Status", "Źródło", "Termin", "Ukończono",
                 "Skuteczność", "Notatki wdrożenia", "Przeterminowane", "Utworzono"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for ci, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    now = datetime.utcnow()
    for ri, a in enumerate(actions, 2):
        org = await s.get(OrgUnit, a.org_unit_id) if a.org_unit_id else None
        is_overdue = a.due_date is not None and a.completed_at is None and a.due_date < now
        ws1.cell(row=ri, column=1, value=f"D-{a.id}")
        ws1.cell(row=ri, column=2, value=a.title)
        ws1.cell(row=ri, column=3, value=a.description or "")
        ws1.cell(row=ri, column=4, value=org.name if org else "")
        ws1.cell(row=ri, column=5, value=a.owner or "")
        ws1.cell(row=ri, column=6, value=a.responsible or "")
        ws1.cell(row=ri, column=7, value=await _de_label(s, a.priority_id) or "")
        ws1.cell(row=ri, column=8, value=await _de_label(s, a.status_id) or "")
        ws1.cell(row=ri, column=9, value=await _de_label(s, a.source_id) or "")
        ws1.cell(row=ri, column=10, value=a.due_date.strftime("%Y-%m-%d") if a.due_date else "")
        ws1.cell(row=ri, column=11, value=a.completed_at.strftime("%Y-%m-%d") if a.completed_at else "")
        ws1.cell(row=ri, column=12, value=f"{a.effectiveness_rating}/5" if a.effectiveness_rating else "")
        ws1.cell(row=ri, column=13, value=a.implementation_notes or "")
        ws1.cell(row=ri, column=14, value="TAK" if is_overdue else "NIE")
        ws1.cell(row=ri, column=15, value=a.created_at.strftime("%Y-%m-%d"))

    # Auto-width
    for col in ws1.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ── Sheet 2: Linked Risks ──
    ws2 = wb.create_sheet("Powiazane ryzyka")
    headers2 = ["Działanie ID", "Działanie tytuł", "Typ", "Obiekt ID", "Nazwa obiektu", "Score", "Poziom ryzyka"]
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
    row2 = 2
    for a in actions:
        links = (await s.execute(select(ActionLink).where(ActionLink.action_id == a.id))).scalars().all()
        for lnk in links:
            ename = await _entity_name(s, lnk.entity_type, lnk.entity_id)
            eextra = await _entity_extra(s, lnk.entity_type, lnk.entity_id)
            ws2.cell(row=row2, column=1, value=f"D-{a.id}")
            ws2.cell(row=row2, column=2, value=a.title)
            ws2.cell(row=row2, column=3, value=lnk.entity_type)
            ws2.cell(row=row2, column=4, value=lnk.entity_id)
            ws2.cell(row=row2, column=5, value=ename or "")
            ws2.cell(row=row2, column=6, value=eextra.get("risk_score") if eextra else None)
            ws2.cell(row=row2, column=7, value=eextra.get("risk_level") if eextra else "")
            row2 += 1

    for col in ws2.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ── Sheet 3: History ──
    ws3 = wb.create_sheet("Historia zmian")
    headers3 = ["Działanie ID", "Pole", "Stara wartość", "Nowa wartość", "Powód", "Data"]
    for ci, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
    row3 = 2
    for a in actions:
        hist = (await s.execute(
            select(ActionHistory).where(ActionHistory.action_id == a.id).order_by(ActionHistory.created_at.desc())
        )).scalars().all()
        for h in hist:
            ws3.cell(row=row3, column=1, value=f"D-{a.id}")
            ws3.cell(row=row3, column=2, value=h.field_name)
            ws3.cell(row=row3, column=3, value=h.old_value or "")
            ws3.cell(row=row3, column=4, value=h.new_value or "")
            ws3.cell(row=row3, column=5, value=h.change_reason or "")
            ws3.cell(row=row3, column=6, value=h.created_at.strftime("%Y-%m-%d %H:%M"))
            row3 += 1

    for col in ws3.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws3.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"dzialania_export_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ═══════════════════ LIST (optimized — batch loading) ═══════════════════

async def _batch_action_outs(s: AsyncSession, actions: list[Action]) -> list[ActionOut]:
    """Build ActionOut list with batch queries instead of N+1."""
    if not actions:
        return []

    now = datetime.utcnow()
    action_ids = [a.id for a in actions]

    # 1) Batch-load all org units
    org_ids = {a.org_unit_id for a in actions if a.org_unit_id}
    org_map: dict[int, str] = {}
    if org_ids:
        rows = (await s.execute(select(OrgUnit.id, OrgUnit.name).where(OrgUnit.id.in_(org_ids)))).all()
        org_map = {r[0]: r[1] for r in rows}

    # 2) Batch-load all dictionary labels
    de_ids = set()
    for a in actions:
        for did in (a.priority_id, a.status_id, a.source_id):
            if did is not None:
                de_ids.add(did)
    de_map: dict[int, str] = {}
    if de_ids:
        rows = (await s.execute(select(DictionaryEntry.id, DictionaryEntry.label).where(DictionaryEntry.id.in_(de_ids)))).all()
        de_map = {r[0]: r[1] for r in rows}

    # 3) Batch-load all links
    links_q = select(ActionLink).where(ActionLink.action_id.in_(action_ids))
    all_links = (await s.execute(links_q)).scalars().all()
    links_by_action: dict[int, list[ActionLink]] = defaultdict(list)
    for lnk in all_links:
        links_by_action[lnk.action_id].append(lnk)

    # 4) Batch-load entity names for linked risks/assets
    risk_ids = {lnk.entity_id for lnk in all_links if lnk.entity_type == "risk"}
    asset_ids = {lnk.entity_id for lnk in all_links if lnk.entity_type == "asset"}
    risk_map: dict[int, Risk] = {}
    asset_name_map: dict[int, str] = {}
    if risk_ids:
        rows = (await s.execute(select(Risk).where(Risk.id.in_(risk_ids)))).scalars().all()
        risk_map = {r.id: r for r in rows}
    if asset_ids:
        rows = (await s.execute(select(Asset.id, Asset.name).where(Asset.id.in_(asset_ids)))).all()
        asset_name_map = {r[0]: r[1] for r in rows}

    # Batch-load risk status labels
    risk_status_ids = {r.status_id for r in risk_map.values() if r.status_id}
    if risk_status_ids - de_ids:
        extra = risk_status_ids - de_ids
        rows = (await s.execute(select(DictionaryEntry.id, DictionaryEntry.label).where(DictionaryEntry.id.in_(extra)))).all()
        for r in rows:
            de_map[r[0]] = r[1]

    # Count other actions linked to same risks (for entity_extra)
    risk_other_count: dict[int, int] = {}
    if risk_ids:
        cnt_q = (
            select(ActionLink.entity_id, func.count(ActionLink.id))
            .where(ActionLink.entity_type == "risk", ActionLink.entity_id.in_(risk_ids))
            .group_by(ActionLink.entity_id)
        )
        rows = (await s.execute(cnt_q)).all()
        risk_other_count = {r[0]: r[1] for r in rows}

    # 5) Batch-load history
    hist_q = (select(ActionHistory)
              .where(ActionHistory.action_id.in_(action_ids))
              .order_by(ActionHistory.created_at.desc()))
    all_history = (await s.execute(hist_q)).scalars().all()
    hist_by_action: dict[int, list[ActionHistory]] = defaultdict(list)
    for h in all_history:
        hist_by_action[h.action_id].append(h)

    # 6) Batch-load attachments
    att_q = (select(ActionAttachment)
             .where(ActionAttachment.action_id.in_(action_ids))
             .order_by(ActionAttachment.created_at.desc()))
    all_atts = (await s.execute(att_q)).scalars().all()
    att_by_action: dict[int, list[ActionAttachment]] = defaultdict(list)
    for att in all_atts:
        att_by_action[att.action_id].append(att)

    # 7) Build results
    results: list[ActionOut] = []
    for action in actions:
        # Links
        link_outs = []
        for lnk in links_by_action.get(action.id, []):
            ename: str | None = None
            eextra: dict | None = None
            if lnk.entity_type == "risk":
                r = risk_map.get(lnk.entity_id)
                if r:
                    ename = r.asset_name
                    eextra = {
                        "risk_score": r.risk_score,
                        "risk_level": r.risk_level,
                        "status_name": de_map.get(r.status_id) if r.status_id else None,
                        "org_unit_name": None,
                        "other_action_count": risk_other_count.get(lnk.entity_id, 0),
                    }
            elif lnk.entity_type == "asset":
                ename = asset_name_map.get(lnk.entity_id)
            else:
                ename = f"{lnk.entity_type}#{lnk.entity_id}"
            link_outs.append(ActionLinkOut(
                id=lnk.id, entity_type=lnk.entity_type,
                entity_id=lnk.entity_id, entity_name=ename,
                entity_extra=eextra, created_at=lnk.created_at,
            ))

        is_overdue = (
            action.due_date is not None
            and action.completed_at is None
            and action.due_date < now
        )

        results.append(ActionOut(
            id=action.id,
            title=action.title,
            description=action.description,
            org_unit_id=action.org_unit_id,
            org_unit_name=org_map.get(action.org_unit_id) if action.org_unit_id else None,
            owner=action.owner,
            responsible=action.responsible,
            priority_id=action.priority_id,
            priority_name=de_map.get(action.priority_id) if action.priority_id else None,
            status_id=action.status_id,
            status_name=de_map.get(action.status_id) if action.status_id else None,
            source_id=action.source_id,
            source_name=de_map.get(action.source_id) if action.source_id else None,
            due_date=action.due_date,
            completed_at=action.completed_at,
            effectiveness_rating=action.effectiveness_rating,
            effectiveness_notes=action.effectiveness_notes,
            implementation_notes=action.implementation_notes,
            is_active=action.is_active,
            is_overdue=is_overdue,
            links=link_outs,
            history=[ActionHistoryOut.model_validate(h) for h in hist_by_action.get(action.id, [])],
            attachments=[ActionAttachmentOut.model_validate(a) for a in att_by_action.get(action.id, [])],
            created_at=action.created_at,
            updated_at=action.updated_at,
        ))

    return results


@router.get("", response_model=list[ActionOut], summary="Lista dzialan")
async def list_actions(
    org_unit_id: int | None = Query(None),
    status_id: int | None = Query(None),
    source_id: int | None = Query(None),
    overdue_only: bool = Query(False),
    include_archived: bool = Query(False),
    entity_type: str | None = Query(None, description="Filter by linked entity type"),
    entity_id: int | None = Query(None, description="Filter by linked entity id"),
    s: AsyncSession = Depends(get_session),
):
    q = select(Action)
    if not include_archived:
        q = q.where(Action.is_active.is_(True))
    if org_unit_id is not None:
        q = q.where(Action.org_unit_id == org_unit_id)
    if status_id is not None:
        q = q.where(Action.status_id == status_id)
    if source_id is not None:
        q = q.where(Action.source_id == source_id)
    if overdue_only:
        q = q.where(Action.due_date < datetime.utcnow()).where(Action.completed_at.is_(None))
    if entity_type and entity_id is not None:
        link_sub = select(ActionLink.action_id).where(
            ActionLink.entity_type == entity_type,
            ActionLink.entity_id == entity_id,
        )
        q = q.where(Action.id.in_(link_sub))
    q = q.order_by(
        case((Action.due_date.is_(None), 1), else_=0),
        Action.due_date.asc(),
        Action.created_at.desc(),
    )
    actions = (await s.execute(q)).scalars().all()
    return await _batch_action_outs(s, actions)


# ═══════════════════ GET ═══════════════════

@router.get("/{action_id}", response_model=ActionOut, summary="Pobierz dzialanie")
async def get_action(action_id: int, s: AsyncSession = Depends(get_session)):
    action = await s.get(Action, action_id)
    if not action:
        raise HTTPException(404, "Dzialanie nie istnieje")
    return await _action_out(s, action)


# ═══════════════════ CREATE ═══════════════════

@router.post("", response_model=ActionOut, status_code=201, summary="Utworz dzialanie")
async def create_action(body: ActionCreate, s: AsyncSession = Depends(get_session)):
    data = body.model_dump(exclude={"links"})
    action = Action(**data)
    s.add(action)
    await s.flush()

    if body.links:
        await _sync_links(s, action.id, body.links)

    await s.commit()
    await s.refresh(action)
    return await _action_out(s, action)


# ═══════════════════ UPDATE ═══════════════════

@router.put("/{action_id}", response_model=ActionOut, summary="Edytuj dzialanie")
async def update_action(action_id: int, body: ActionUpdate, s: AsyncSession = Depends(get_session)):
    action = await s.get(Action, action_id)
    if not action:
        raise HTTPException(404, "Dzialanie nie istnieje")

    reason = body.change_reason
    data = body.model_dump(exclude_unset=True, exclude={"links", "change_reason"})
    for k, v in data.items():
        old = getattr(action, k)
        if old != v:
            await _track_change(s, action_id, k, old, v, change_reason=reason)
            setattr(action, k, v)

    if body.links is not None:
        await _sync_links(s, action_id, body.links)

    await s.commit()
    await s.refresh(action)
    return await _action_out(s, action)


# ═══════════════════ CLOSE (with effectiveness) ═══════════════════

@router.post("/{action_id}/close", response_model=ActionOut, summary="Zamknij dzialanie z ocena skutecznosci")
async def close_action(action_id: int, body: ActionCloseRequest, s: AsyncSession = Depends(get_session)):
    action = await s.get(Action, action_id)
    if not action:
        raise HTTPException(404, "Dzialanie nie istnieje")

    reason = body.change_reason or "Zamkniecie dzialania"

    await _track_change(s, action_id, "completed_at", action.completed_at, datetime.utcnow(), change_reason=reason)
    action.completed_at = datetime.utcnow()
    action.effectiveness_rating = body.effectiveness_rating
    action.effectiveness_notes = body.effectiveness_notes
    if body.implementation_notes is not None:
        await _track_change(s, action_id, "implementation_notes", action.implementation_notes, body.implementation_notes, change_reason=reason)
        action.implementation_notes = body.implementation_notes

    # Find "completed" status
    from app.models.dictionary import DictionaryType
    q = (
        select(DictionaryEntry.id)
        .select_from(DictionaryEntry)
        .join(DictionaryType, DictionaryEntry.dict_type_id == DictionaryType.id)
        .where(DictionaryType.code == "action_status")
        .where(DictionaryEntry.code == "completed")
        .limit(1)
    )
    completed_id = (await s.execute(q)).scalar()
    if completed_id:
        await _track_change(s, action_id, "status_id", action.status_id, completed_id, change_reason=reason)
        action.status_id = completed_id

    await s.commit()
    await s.refresh(action)
    return await _action_out(s, action)


# ═══════════════════ ARCHIVE ═══════════════════

@router.delete("/{action_id}", summary="Archiwizuj dzialanie")
async def archive_action(action_id: int, s: AsyncSession = Depends(get_session)):
    action = await s.get(Action, action_id)
    if not action:
        raise HTTPException(404, "Dzialanie nie istnieje")
    action.is_active = False
    await s.commit()
    await s.refresh(action)
    return {"status": "archived", "id": action_id}


# ═══════════════════ ATTACHMENTS ═══════════════════

@router.post("/{action_id}/attachments", response_model=ActionAttachmentOut, summary="Dodaj zalacznik")
async def upload_attachment(
    action_id: int,
    file: UploadFile = File(...),
    s: AsyncSession = Depends(get_session),
):
    action = await s.get(Action, action_id)
    if not action:
        raise HTTPException(404, "Dzialanie nie istnieje")

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    ext = os.path.splitext(file.filename or "file")[1]
    stored_name = f"{action_id}_{uuid.uuid4().hex}{ext}"
    path = os.path.join(UPLOAD_DIR, stored_name)

    content = await file.read()
    with open(path, "wb") as f:
        f.write(content)

    att = ActionAttachment(
        action_id=action_id,
        filename=stored_name,
        original_name=file.filename or "file",
        file_size=len(content),
        content_type=file.content_type,
    )
    s.add(att)
    await s.commit()
    await s.refresh(att)
    return ActionAttachmentOut.model_validate(att)


@router.get("/{action_id}/attachments/{attachment_id}/download", summary="Pobierz zalacznik")
async def download_attachment(
    action_id: int,
    attachment_id: int,
    s: AsyncSession = Depends(get_session),
):
    att = await s.get(ActionAttachment, attachment_id)
    if not att or att.action_id != action_id:
        raise HTTPException(404, "Zalacznik nie istnieje")

    path = os.path.join(UPLOAD_DIR, att.filename)
    if not os.path.exists(path):
        raise HTTPException(404, "Plik nie istnieje na dysku")

    return FileResponse(path, filename=att.original_name, media_type=att.content_type or "application/octet-stream")


@router.delete("/{action_id}/attachments/{attachment_id}", summary="Usun zalacznik")
async def delete_attachment(
    action_id: int,
    attachment_id: int,
    s: AsyncSession = Depends(get_session),
):
    att = await s.get(ActionAttachment, attachment_id)
    if not att or att.action_id != action_id:
        raise HTTPException(404, "Zalacznik nie istnieje")

    path = os.path.join(UPLOAD_DIR, att.filename)
    if os.path.exists(path):
        os.remove(path)

    await s.delete(att)
    await s.commit()
    return {"status": "deleted", "id": attachment_id}


# ═══════════════════ COMMENTS ═══════════════════

@router.get("/{action_id}/comments", response_model=list[ActionCommentOut], summary="Lista komentarzy")
async def list_comments(action_id: int, s: AsyncSession = Depends(get_session)):
    action = await s.get(Action, action_id)
    if not action:
        raise HTTPException(404, "Dzialanie nie istnieje")
    q = (select(ActionComment)
         .where(ActionComment.action_id == action_id)
         .order_by(ActionComment.created_at.asc()))
    comments = (await s.execute(q)).scalars().all()
    return [ActionCommentOut.model_validate(c) for c in comments]


@router.post("/{action_id}/comments", response_model=ActionCommentOut, status_code=201, summary="Dodaj komentarz")
async def add_comment(action_id: int, body: ActionCommentCreate, s: AsyncSession = Depends(get_session)):
    action = await s.get(Action, action_id)
    if not action:
        raise HTTPException(404, "Dzialanie nie istnieje")
    comment = ActionComment(
        action_id=action_id,
        author=body.author,
        content=body.content,
    )
    s.add(comment)
    await s.commit()
    await s.refresh(comment)
    return ActionCommentOut.model_validate(comment)


@router.put("/{action_id}/comments/{comment_id}", response_model=ActionCommentOut, summary="Edytuj komentarz")
async def update_comment(action_id: int, comment_id: int, body: ActionCommentUpdate, s: AsyncSession = Depends(get_session)):
    comment = await s.get(ActionComment, comment_id)
    if not comment or comment.action_id != action_id:
        raise HTTPException(404, "Komentarz nie istnieje")
    comment.content = body.content
    await s.commit()
    await s.refresh(comment)
    return ActionCommentOut.model_validate(comment)


@router.delete("/{action_id}/comments/{comment_id}", summary="Usun komentarz")
async def delete_comment(action_id: int, comment_id: int, s: AsyncSession = Depends(get_session)):
    comment = await s.get(ActionComment, comment_id)
    if not comment or comment.action_id != action_id:
        raise HTTPException(404, "Komentarz nie istnieje")
    await s.delete(comment)
    await s.commit()
    return {"status": "deleted", "id": comment_id}
