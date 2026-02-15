"""
Report generation — /api/v1/reports
Generates Excel reports for risks, assets, assessments, and executive summary.
"""
import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy import case, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_session
from app.models.asset import Asset, AssetRelationship
from app.models.asset_category import AssetCategory
from app.models.dictionary import DictionaryEntry
from app.models.framework import Assessment, Framework
from app.models.incident import Incident
from app.models.org_unit import OrgUnit
from app.models.risk import Risk
from app.models.security_area import SecurityDomain
from app.models.vendor import Vendor
from app.models.vulnerability import VulnerabilityRecord

router = APIRouter(prefix="/api/v1/reports", tags=["Raporty"])


def _wb():
    """Create openpyxl workbook (lazy import)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        return Workbook(), Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        return Workbook(), Font, PatternFill, Alignment, Border, Side


def _styled_header(ws, row, headers, Font, PatternFill, Alignment, Border, Side):
    """Apply styled headers to a worksheet row."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    return thin_border


def _stream_xlsx(wb) -> StreamingResponse:
    """Stream workbook as xlsx download."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=raport_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"},
    )


async def _de_label(s: AsyncSession, entry_id: int | None) -> str | None:
    if entry_id is None:
        return None
    e = await s.get(DictionaryEntry, entry_id)
    return e.label if e else None


# ═══════════════════ RISK REPORT ═══════════════════

@router.get("/risks", summary="Raport ryzyk (Excel)")
async def report_risks(
    org_unit_id: int | None = Query(None),
    s: AsyncSession = Depends(get_session),
):
    wb, Font, PatternFill, Alignment, Border, Side = _wb()
    ws = wb.active
    ws.title = "Rejestr Ryzyk"

    headers = [
        "ID", "Aktywo", "Jednostka org.", "Domena bezp.", "Kategoria ryzyka",
        "Wplyw (W)", "Prawdop. (P)", "Zabezp. (Z)", "Ocena ryzyka (R)", "Poziom",
        "Status", "Strategia", "Wlasciciel", "Termin realizacji",
        "Ryzyko rezydualne", "Zaakceptowal", "Data akceptacji",
    ]
    border = _styled_header(ws, 1, headers, Font, PatternFill, Alignment, Border, Side)

    q = select(Risk).where(Risk.is_active.is_(True))
    if org_unit_id:
        q = q.where(Risk.org_unit_id == org_unit_id)
    q = q.order_by(Risk.risk_score.desc())
    risks = (await s.execute(q)).scalars().all()

    for row_idx, r in enumerate(risks, 2):
        org = await s.get(OrgUnit, r.org_unit_id) if r.org_unit_id else None
        area = await s.get(SecurityDomain, r.security_area_id) if r.security_area_id else None
        values = [
            f"R-{r.id}", r.asset_name, org.name if org else "",
            area.name if area else "", await _de_label(s, r.risk_category_id),
            r.impact_level, r.probability_level, float(r.safeguard_rating),
            float(r.risk_score) if r.risk_score else 0, r.risk_level,
            await _de_label(s, r.status_id), await _de_label(s, r.strategy_id),
            r.owner, r.treatment_deadline.isoformat() if r.treatment_deadline else "",
            float(r.residual_risk) if r.residual_risk else "", r.accepted_by or "",
            r.accepted_at.isoformat() if r.accepted_at else "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    return _stream_xlsx(wb)


# ═══════════════════ ASSET REPORT ═══════════════════

@router.get("/assets", summary="Raport aktywow (Excel)")
async def report_assets(
    asset_category_id: int | None = Query(None),
    s: AsyncSession = Depends(get_session),
):
    wb, Font, PatternFill, Alignment, Border, Side = _wb()
    ws = wb.active
    ws.title = "Rejestr Aktywow"

    headers = [
        "ID", "Ref", "Nazwa", "Kategoria CMDB", "Jednostka org.",
        "Wlasciciel", "Lokalizacja", "Wrazliwosc", "Krytycznosc",
        "Ilosc ryzyk", "Utworzono",
    ]
    border = _styled_header(ws, 1, headers, Font, PatternFill, Alignment, Border, Side)

    q = select(Asset).where(Asset.is_active.is_(True))
    if asset_category_id:
        q = q.where(Asset.asset_category_id == asset_category_id)
    q = q.order_by(Asset.name)
    assets = (await s.execute(q)).scalars().all()

    for row_idx, a in enumerate(assets, 2):
        org = await s.get(OrgUnit, a.org_unit_id) if a.org_unit_id else None
        cat = await s.get(AssetCategory, a.asset_category_id) if a.asset_category_id else None
        rc_q = select(func.count()).select_from(Risk).where(Risk.asset_id == a.id)
        rc = (await s.execute(rc_q)).scalar() or 0
        values = [
            a.id, a.ref_id or "", a.name,
            cat.name if cat else "", org.name if org else "",
            a.owner or "", a.location or "",
            await _de_label(s, a.sensitivity_id), await _de_label(s, a.criticality_id),
            rc, a.created_at.strftime("%Y-%m-%d"),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    return _stream_xlsx(wb)


# ═══════════════════ EXECUTIVE REPORT ═══════════════════

@router.get("/executive", summary="Raport Executive Summary (Excel)")
async def report_executive(s: AsyncSession = Depends(get_session)):
    wb, Font, PatternFill, Alignment, Border, Side = _wb()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Podsumowanie"
    title_font = Font(bold=True, size=14, color="1F4E79")
    ws1.cell(row=1, column=1, value="SecurePosture — Raport Executive Summary").font = title_font
    ws1.cell(row=2, column=1, value=f"Data generacji: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    # KPIs
    ws1.cell(row=4, column=1, value="Kluczowe wskazniki (KPI)").font = Font(bold=True, size=12)

    # Risk counts
    total_risks = (await s.execute(select(func.count()).select_from(Risk).where(Risk.is_active.is_(True)))).scalar() or 0
    high_risks = (await s.execute(select(func.count()).select_from(Risk).where(Risk.is_active.is_(True), Risk.risk_level == "high"))).scalar() or 0
    medium_risks = (await s.execute(select(func.count()).select_from(Risk).where(Risk.is_active.is_(True), Risk.risk_level == "medium"))).scalar() or 0
    low_risks = (await s.execute(select(func.count()).select_from(Risk).where(Risk.is_active.is_(True), Risk.risk_level == "low"))).scalar() or 0

    # Asset counts
    total_assets = (await s.execute(select(func.count()).select_from(Asset).where(Asset.is_active.is_(True)))).scalar() or 0
    assets_with_risks = (await s.execute(
        select(func.count(func.distinct(Risk.asset_id))).where(Risk.is_active.is_(True), Risk.asset_id.isnot(None))
    )).scalar() or 0

    # Vulnerability counts
    open_vulns = (await s.execute(select(func.count()).select_from(VulnerabilityRecord).where(VulnerabilityRecord.is_active.is_(True)))).scalar() or 0

    # Incident counts
    open_incidents = (await s.execute(select(func.count()).select_from(Incident).where(Incident.is_active.is_(True)))).scalar() or 0

    kpis = [
        ("Ryzyka ogolnie", total_risks),
        ("Ryzyka wysokie", high_risks),
        ("Ryzyka srednie", medium_risks),
        ("Ryzyka niskie", low_risks),
        ("Aktywow ogolnie", total_assets),
        ("Aktywa z ryzykami", assets_with_risks),
        ("Otwarte podatnosci", open_vulns),
        ("Otwarte incydenty", open_incidents),
    ]
    for idx, (label, val) in enumerate(kpis):
        ws1.cell(row=6 + idx, column=1, value=label)
        ws1.cell(row=6 + idx, column=2, value=val).font = Font(bold=True)

    # Sheet 2: Risk summary
    ws2 = wb.create_sheet("Ryzyka - Top 20")
    headers = ["ID", "Aktywo", "Jednostka", "Domena", "Ocena (R)", "Poziom", "Status", "Wlasciciel"]
    border = _styled_header(ws2, 1, headers, Font, PatternFill, Alignment, Border, Side)
    top_q = select(Risk).where(Risk.is_active.is_(True)).order_by(Risk.risk_score.desc()).limit(20)
    top_risks = (await s.execute(top_q)).scalars().all()
    for i, r in enumerate(top_risks, 2):
        org = await s.get(OrgUnit, r.org_unit_id) if r.org_unit_id else None
        area = await s.get(SecurityDomain, r.security_area_id) if r.security_area_id else None
        vals = [f"R-{r.id}", r.asset_name, org.name if org else "", area.name if area else "",
                float(r.risk_score) if r.risk_score else 0, r.risk_level,
                await _de_label(s, r.status_id), r.owner or ""]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=i, column=c, value=v).border = border

    for col in ws2.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    # Sheet 3: Assets overview
    ws3 = wb.create_sheet("Aktywa - Przeglad")
    headers3 = ["Kategoria CMDB", "Ilosc aktywow"]
    _styled_header(ws3, 1, headers3, Font, PatternFill, Alignment, Border, Side)
    cat_q = (
        select(AssetCategory.name, func.count(Asset.id))
        .join(Asset, Asset.asset_category_id == AssetCategory.id, isouter=True)
        .where(AssetCategory.is_active.is_(True))
        .group_by(AssetCategory.id, AssetCategory.name)
        .order_by(func.count(Asset.id).desc())
    )
    cat_rows = (await s.execute(cat_q)).all()
    for i, (name, count) in enumerate(cat_rows, 2):
        ws3.cell(row=i, column=1, value=name)
        ws3.cell(row=i, column=2, value=count)

    return _stream_xlsx(wb)


# ═══════════════════ AI MANAGEMENT REPORT ═══════════════════

@router.get("/ai-management", summary="Raport zarzadczy AI")
async def report_ai_management(
    org_unit_id: int | None = Query(None),
    s: AsyncSession = Depends(get_session),
):
    """Generate AI-powered management report from current database state."""
    import logging
    from sqlalchemy import text as sa_text

    from app.models.security_area import SecurityDomain
    from app.services.ai_service import (
        AIFeatureDisabledException,
        AINotConfiguredException,
        AIParsingError,
        AIRateLimitException,
        get_ai_service,
    )

    log = logging.getLogger(__name__)

    try:
        ai = await get_ai_service(s)
    except Exception as exc:
        log.warning("get_ai_service failed: %s", exc)
        raise HTTPException(
            status_code=422,
            detail="AI nie jest skonfigurowane. Przejdz do Ustawienia > Integracja AI.",
        )
    if not ai.is_available:
        raise HTTPException(
            status_code=422,
            detail="AI nie jest skonfigurowane. Przejdz do Ustawienia > Integracja AI.",
        )

    # ── Gather all data (each section wrapped so partial failures don't crash) ──

    flt = Risk.org_unit_id == org_unit_id if org_unit_id else True

    # Risk counts by level
    try:
        level_q = (
            select(Risk.risk_level, func.count().label("cnt"))
            .where(Risk.is_active.is_(True))
            .where(flt)
            .group_by(Risk.risk_level)
        )
        level_rows = (await s.execute(level_q)).all()
        risk_counts: dict[str | None, int] = {r.risk_level: r.cnt for r in level_rows}
        total_risks = sum(risk_counts.values())
    except Exception:
        risk_counts, total_risks = {}, 0

    # Average risk score
    try:
        avg_score = (await s.execute(
            select(func.avg(Risk.risk_score)).where(Risk.is_active.is_(True)).where(flt)
        )).scalar()
    except Exception:
        avg_score = None

    # Top 10 risks
    try:
        top_q = (
            select(Risk.id, Risk.asset_name, Risk.risk_score, Risk.risk_level, Risk.owner)
            .where(Risk.is_active.is_(True)).where(flt)
            .order_by(Risk.risk_score.desc()).limit(10)
        )
        top_risks = (await s.execute(top_q)).all()
        top_risks_text = "\n".join(
            f"  - R-{r.id}: {r.asset_name}, score={float(r.risk_score):.1f}, "
            f"poziom={r.risk_level}, wlasciciel={r.owner or 'brak'}"
            for r in top_risks
        )
    except Exception:
        top_risks_text = "  Brak danych."

    # Risks by security domain
    try:
        area_q = (
            select(SecurityDomain.name, func.count(Risk.id).label("cnt"))
            .join(Risk, Risk.security_area_id == SecurityDomain.id, isouter=True)
            .where(SecurityDomain.is_active.is_(True))
            .group_by(SecurityDomain.name)
            .order_by(func.count(Risk.id).desc())
        )
        area_rows = (await s.execute(area_q)).all()
        risks_by_area_text = "\n".join(
            f"  - {r.name}: {r.cnt} ryzyk" for r in area_rows if r.cnt
        ) or "  Brak danych."
    except Exception:
        risks_by_area_text = "  Brak danych."

    # Risks by org unit
    try:
        org_q = (
            select(OrgUnit.name, func.count(Risk.id).label("cnt"))
            .join(Risk, Risk.org_unit_id == OrgUnit.id, isouter=True)
            .where(OrgUnit.is_active.is_(True))
            .group_by(OrgUnit.name)
        )
        org_rows = (await s.execute(org_q)).all()
        risks_by_org_text = "\n".join(
            f"  - {r.name}: {r.cnt} ryzyk" for r in org_rows if r.cnt
        ) or "  Brak danych."
    except Exception:
        risks_by_org_text = "  Brak danych."

    # Overdue risks (TIMESTAMPDIFF for MySQL, fallback for others)
    interval = 90
    overdue_count = 0
    try:
        from app.models.risk import RiskReviewConfig
        cfg_row = (await s.execute(select(RiskReviewConfig.review_interval_days))).scalar()
        interval = cfg_row or 90
    except Exception:
        pass

    try:
        overdue_q = (
            select(func.count())
            .select_from(Risk)
            .where(
                Risk.is_active.is_(True),
                flt,
                func.timestampdiff(
                    sa_text("DAY"),
                    func.coalesce(Risk.last_review_at, Risk.identified_at),
                    func.now(),
                ) > interval,
            )
        )
        overdue_count = (await s.execute(overdue_q)).scalar() or 0
    except Exception:
        overdue_count = 0

    # Asset counts
    try:
        total_assets = (await s.execute(
            select(func.count()).select_from(Asset).where(Asset.is_active.is_(True))
        )).scalar() or 0
    except Exception:
        total_assets = 0

    # Assets by category
    try:
        asset_cat_q = (
            select(AssetCategory.name, func.count(Asset.id).label("cnt"))
            .join(Asset, Asset.asset_category_id == AssetCategory.id, isouter=True)
            .where(AssetCategory.is_active.is_(True))
            .group_by(AssetCategory.name)
            .order_by(func.count(Asset.id).desc())
            .limit(15)
        )
        asset_cat_rows = (await s.execute(asset_cat_q)).all()
        assets_by_cat_text = "\n".join(
            f"  - {r.name}: {r.cnt}" for r in asset_cat_rows if r.cnt
        ) or "  Brak danych."
    except Exception:
        assets_by_cat_text = "  Brak danych."

    # Vulnerability and incident counts
    try:
        open_vulns = (await s.execute(
            select(func.count()).select_from(VulnerabilityRecord).where(VulnerabilityRecord.is_active.is_(True))
        )).scalar() or 0
    except Exception:
        open_vulns = 0

    try:
        open_incidents = (await s.execute(
            select(func.count()).select_from(Incident).where(Incident.is_active.is_(True))
        )).scalar() or 0
    except Exception:
        open_incidents = 0

    # Framework assessments
    try:
        fw_q = (
            select(Framework.name, Assessment.overall_score, Assessment.completion_pct, Assessment.status)
            .join(Assessment, Assessment.framework_id == Framework.id)
            .where(Assessment.is_active.is_(True))
            .order_by(Assessment.created_at.desc())
            .limit(5)
        )
        fw_rows = (await s.execute(fw_q)).all()
        fw_text = "\n".join(
            f"  - {r.name}: score={float(r.overall_score):.1f}/100, "
            f"kompletnosc={float(r.completion_pct):.0f}%, status={r.status}"
            for r in fw_rows if r.overall_score
        ) or "  Brak ocen frameworkow."
    except Exception:
        fw_text = "  Brak ocen frameworkow."

    # Security posture score
    try:
        from app.services.dashboard import get_posture_score
        posture = await get_posture_score(s, org_unit_id)
        posture_text = (
            f"  Ogolny wynik: {posture.score}/100 (ocena: {posture.grade})\n"
            + "\n".join(
                f"  - {d.name}: {d.score}/100 (waga: {d.weight*100:.0f}%)"
                for d in posture.dimensions
            )
        )
    except Exception:
        posture_text = "  Brak danych Security Posture Score."

    # ── Build LLM context ──

    data_context = f"""=== DANE ORGANIZACJI — stan na {datetime.now().strftime('%Y-%m-%d %H:%M')} ===

--- RYZYKA ---
Laczna liczba aktywnych ryzyk: {total_risks}
Rozklad wg poziomu: wysokie={risk_counts.get('high', 0)}, srednie={risk_counts.get('medium', 0)}, niskie={risk_counts.get('low', 0)}
Sredni score ryzyka: {float(avg_score):.1f if avg_score else 'brak danych'}
Ryzyka przeterminowane (brak przegladu > {interval} dni): {overdue_count}

Top 10 ryzyk wg score:
{top_risks_text}

Ryzyka wg domeny bezpieczenstwa:
{risks_by_area_text}

Ryzyka wg jednostki organizacyjnej:
{risks_by_org_text}

--- AKTYWA (CMDB) ---
Laczna liczba aktywow: {total_assets}
Rozklad wg kategorii CMDB:
{assets_by_cat_text}

--- PODATNOSCI I INCYDENTY ---
Otwarte podatnosci: {open_vulns}
Otwarte incydenty: {open_incidents}

--- OCENY FRAMEWORKOW (compliance) ---
{fw_text}

--- SECURITY POSTURE SCORE ---
{posture_text}

Wygeneruj profesjonalny raport zarzadczy na podstawie powyzszych danych."""

    # ── Call AI ──
    try:
        result = await ai.generate_management_report(user_id=1, data_context=data_context)
    except AINotConfiguredException as e:
        raise HTTPException(status_code=422, detail=str(e))
    except AIFeatureDisabledException as e:
        raise HTTPException(status_code=422, detail=str(e))
    except AIRateLimitException as e:
        raise HTTPException(status_code=429, detail=str(e))
    except AIParsingError as e:
        raise HTTPException(status_code=502, detail=f"AI zwrocilo nieprawidlowa odpowiedz: {e}")
    except Exception as e:
        log.exception("AI management report generation failed")
        raise HTTPException(status_code=502, detail=f"Blad komunikacji z AI: {e}")

    return JSONResponse(content={
        "generated_at": datetime.now().isoformat(),
        "report": result,
    })
