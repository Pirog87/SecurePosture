"""
Framework Import Service — parses CISO Assistant Excel/YAML files.

Supports:
  - Excel v2 format (tabs: library_content + framework nodes)
  - YAML native format (flat list with parent_urn + depth, or nested children)
  - Default dimension creation when framework has no scale defined

Key fix (v2): URN normalization to lowercase + two-pass parent resolution
to handle CISO Assistant files where parent_urn references are not always
ordered before their children.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from decimal import Decimal
from typing import Any, BinaryIO

import yaml
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.framework import (
    AssessmentDimension, DimensionLevel, Framework, FrameworkNode,
    FrameworkVersionHistory,
)

log = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# Default scale (used when framework has none)
# ─────────────────────────────────────────────

DEFAULT_DIMENSION = {
    "dimension_key": "compliance_level",
    "name": "Compliance Level",
    "name_pl": "Poziom zgodnosci",
    "levels": [
        {"order": 0, "value": 0.00, "label": "Not implemented", "label_pl": "Niezaimplementowane", "color": "#EF4444"},
        {"order": 1, "value": 0.33, "label": "Partially implemented", "label_pl": "Czesciowo", "color": "#EAB308"},
        {"order": 2, "value": 0.66, "label": "Largely implemented", "label_pl": "W duzej mierze", "color": "#22C55E"},
        {"order": 3, "value": 1.00, "label": "Fully implemented", "label_pl": "W pelni", "color": "#16A34A"},
    ],
}


def _normalize_urn(urn: str | None) -> str | None:
    """Normalize URN to lowercase and strip whitespace for consistent lookups."""
    if not urn:
        return None
    return urn.strip().lower()


async def _create_default_dimensions(s: AsyncSession, fw: Framework) -> int:
    """Create the default single-dimension scale for a framework."""
    d = DEFAULT_DIMENSION
    dim = AssessmentDimension(
        framework_id=fw.id,
        dimension_key=d["dimension_key"],
        name=d["name"],
        name_pl=d["name_pl"],
        order_id=1,
        weight=Decimal("1.00"),
    )
    s.add(dim)
    await s.flush()

    for lvl_data in d["levels"]:
        s.add(DimensionLevel(
            dimension_id=dim.id,
            level_order=lvl_data["order"],
            value=Decimal(str(lvl_data["value"])),
            label=lvl_data["label"],
            label_pl=lvl_data["label_pl"],
            color=lvl_data["color"],
        ))

    return 1  # 1 dimension created


async def _create_dimensions_from_score_def(
    s: AsyncSession, fw: Framework, score_def: list[dict]
) -> int:
    """Create dimensions + levels from a CISO Assistant score definition."""
    count = 0
    for idx, dim_def in enumerate(score_def):
        dim = AssessmentDimension(
            framework_id=fw.id,
            dimension_key=dim_def.get("key", f"dim_{idx+1}"),
            name=dim_def.get("name", f"Dimension {idx+1}"),
            name_pl=dim_def.get("name_pl"),
            order_id=idx + 1,
            weight=Decimal(str(dim_def.get("weight", 1.0))),
        )
        s.add(dim)
        await s.flush()

        for lvl_idx, lvl_data in enumerate(dim_def.get("levels", [])):
            s.add(DimensionLevel(
                dimension_id=dim.id,
                level_order=lvl_idx,
                value=Decimal(str(lvl_data.get("value", 0))),
                label=lvl_data.get("label", ""),
                label_pl=lvl_data.get("label_pl"),
                color=lvl_data.get("color"),
            ))
        count += 1
    return count


async def _create_dimensions_from_scores_definition(
    s: AsyncSession, fw: Framework, scores_def: list[dict]
) -> int:
    """Create dimensions from CISO Assistant scores_definition format.

    CISO Assistant uses 'scores_definition' with score/name/description fields
    to define maturity levels as a single dimension.
    """
    dim = AssessmentDimension(
        framework_id=fw.id,
        dimension_key="maturity",
        name="Maturity Level",
        name_pl="Poziom dojrzalosci",
        order_id=1,
        weight=Decimal("1.00"),
    )
    s.add(dim)
    await s.flush()

    for idx, score_item in enumerate(scores_def):
        score_val = score_item.get("score", idx)
        max_score = max((s_item.get("score", 0) for s_item in scores_def), default=1) or 1
        normalized = float(score_val) / float(max_score)

        colors = ["#EF4444", "#F97316", "#EAB308", "#22C55E", "#16A34A", "#059669"]
        color = colors[min(idx, len(colors) - 1)]

        s.add(DimensionLevel(
            dimension_id=dim.id,
            level_order=idx,
            value=Decimal(str(round(normalized, 2))),
            label=score_item.get("name", f"Level {score_val}"),
            label_pl=None,
            description=score_item.get("description"),
            color=color,
        ))

    return 1


# ===================================================
# EXCEL PARSER (CISO Assistant v2 format)
# ===================================================

async def import_from_excel(
    s: AsyncSession, file: BinaryIO, imported_by: str = "admin"
) -> Framework:
    """Import framework from CISO Assistant Excel v2 format."""
    wb = load_workbook(file, read_only=True, data_only=True)

    # 1. Read library_content metadata tab
    meta = _read_excel_metadata(wb)

    # 2. Check for duplicate URN
    fw_urn = _normalize_urn(meta.get("framework_urn"))
    if fw_urn:
        existing = (await s.execute(
            select(Framework).where(Framework.urn == fw_urn)
        )).scalar_one_or_none()
        if existing:
            raise ValueError(f"Framework with URN '{fw_urn}' already exists (id={existing.id})")

    # 3. Create Framework record
    fw = Framework(
        urn=fw_urn,
        ref_id=meta.get("framework_ref_id") or meta.get("ref_id"),
        name=meta.get("framework_name") or meta.get("library_name", "Unnamed Framework"),
        description=meta.get("framework_description") or meta.get("library_description"),
        version=meta.get("library_version"),
        provider=meta.get("library_provider"),
        packager=meta.get("library_packager"),
        copyright=meta.get("library_copyright"),
        source_format="ciso_assistant_excel",
        locale=meta.get("library_locale", "en"),
        implementation_groups_definition=meta.get("implementation_groups_definition"),
        imported_at=datetime.utcnow(),
        imported_by=imported_by,
        lifecycle_status="published",
        edit_version=1,
        published_version=meta.get("library_version"),
    )
    s.add(fw)
    await s.flush()

    # 4. Read nodes from framework tab
    tab_name = meta.get("framework_ref_id") or meta.get("ref_id")
    nodes_data = _read_excel_nodes(wb, tab_name)

    # 5. Insert nodes
    total, assessable = await _insert_nodes(s, fw, nodes_data)
    fw.total_nodes = total
    fw.total_assessable = assessable

    # 6. Create dimensions
    score_def = meta.get("score_definition")
    if score_def and isinstance(score_def, list):
        dims_count = await _create_dimensions_from_score_def(s, fw, score_def)
    else:
        dims_count = await _create_default_dimensions(s, fw)

    # 7. Create initial version record
    s.add(FrameworkVersionHistory(
        framework_id=fw.id,
        edit_version=1,
        lifecycle_status="published",
        change_summary=f"Import z Excel: {fw.name}",
        changed_by=imported_by,
        snapshot_nodes_count=total,
        snapshot_assessable_count=assessable,
    ))

    wb.close()
    return fw


def _read_excel_metadata(wb) -> dict[str, Any]:
    """Read the library_content tab for metadata."""
    meta = {}

    if "library_content" not in wb.sheetnames:
        # Try first sheet as fallback
        return meta

    ws = wb["library_content"]

    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] and row[1]:
            key = str(row[0]).strip().lower()
            val = row[1]
            # Normalize keys
            key = key.replace(" ", "_")
            meta[key] = val

    # Parse JSON fields
    for json_key in ("implementation_groups_definition", "score_definition"):
        if json_key in meta and isinstance(meta[json_key], str):
            import json
            try:
                meta[json_key] = json.loads(meta[json_key])
            except (json.JSONDecodeError, TypeError):
                pass

    return meta


def _read_excel_nodes(wb, tab_name: str | None) -> list[dict]:
    """Read the framework nodes from the named tab."""
    # Find the right tab
    ws = None
    if tab_name and tab_name in wb.sheetnames:
        ws = wb[tab_name]
    else:
        # Try tabs that aren't 'library_content'
        for name in wb.sheetnames:
            if name.lower() != "library_content":
                ws = wb[name]
                break

    if ws is None:
        return []

    # Read header row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers.append(str(cell).strip().lower() if cell else "")

    nodes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = val
        if row_dict.get("ref_id") or row_dict.get("name"):
            nodes.append(row_dict)

    return nodes


# ===================================================
# YAML PARSER (CISO Assistant native format)
# ===================================================

def _extract_framework_from_yaml(data: dict) -> tuple[dict, dict]:
    """Extract the framework dict and the top-level library metadata from YAML.

    Handles multiple CISO Assistant YAML layouts:
      1. objects.framework  (dict -- current format)
      2. objects.frameworks (list)
      3. objects as list of typed objects (legacy)
      4. Direct framework at top level (no objects wrapper)

    Returns (fw_data, library_meta) where library_meta is the top-level dict.
    """
    if "objects" not in data:
        return data, data

    objects = data["objects"]

    # Current format: objects is a dict with key "framework" (singular)
    if isinstance(objects, dict):
        if "framework" in objects:
            fw = objects["framework"]
            if isinstance(fw, dict):
                return fw, data
        if "frameworks" in objects:
            fws = objects["frameworks"]
            if isinstance(fws, list) and fws:
                return fws[0], data
        # Fallback: first dict value that has requirement_nodes
        for val in objects.values():
            if isinstance(val, dict) and "requirement_nodes" in val:
                return val, data
            if isinstance(val, list):
                for item in val:
                    if isinstance(item, dict) and "requirement_nodes" in item:
                        return item, data
        return data, data

    # Legacy format: objects is a list of typed objects
    if isinstance(objects, list):
        for obj in objects:
            if isinstance(obj, dict) and obj.get("type") == "framework":
                return obj, data
        # Fallback: first item with requirement_nodes
        for obj in objects:
            if isinstance(obj, dict) and "requirement_nodes" in obj:
                return obj, data

    return data, data


async def import_from_yaml(
    s: AsyncSession, file: BinaryIO, imported_by: str = "admin"
) -> Framework:
    """Import framework from CISO Assistant YAML format."""
    content = file.read()
    if isinstance(content, bytes):
        content = content.decode("utf-8")
    data = yaml.safe_load(content)

    if not data:
        raise ValueError("Empty YAML file")

    fw_data, lib_meta = _extract_framework_from_yaml(data)

    # Check for duplicate URN
    fw_urn = _normalize_urn(fw_data.get("urn") or lib_meta.get("urn"))
    if fw_urn:
        existing = (await s.execute(
            select(Framework).where(Framework.urn == fw_urn)
        )).scalar_one_or_none()
        if existing:
            raise ValueError(f"Framework with URN '{fw_urn}' already exists (id={existing.id})")

    # version can be int in YAML -- convert to str
    version_raw = fw_data.get("version") or lib_meta.get("version")
    version_str = str(version_raw) if version_raw is not None else None

    # Create Framework record
    fw = Framework(
        urn=fw_urn,
        ref_id=fw_data.get("ref_id") or lib_meta.get("ref_id"),
        name=fw_data.get("name") or lib_meta.get("name", "Unnamed Framework"),
        description=fw_data.get("description") or lib_meta.get("description"),
        version=version_str,
        provider=fw_data.get("provider") or lib_meta.get("provider"),
        packager=fw_data.get("packager") or lib_meta.get("packager"),
        copyright=fw_data.get("copyright") or lib_meta.get("copyright"),
        source_format="ciso_assistant_yaml",
        locale=fw_data.get("locale") or lib_meta.get("locale", "en"),
        implementation_groups_definition=fw_data.get("implementation_groups_definition"),
        imported_at=datetime.utcnow(),
        imported_by=imported_by,
        lifecycle_status="published",
        edit_version=1,
        published_version=version_str,
    )
    s.add(fw)
    await s.flush()

    # Parse nodes -- CISO Assistant uses flat list with parent_urn + depth
    req_nodes = fw_data.get("requirement_nodes", [])
    if not req_nodes:
        log.warning("No requirement_nodes found in YAML for framework %s", fw.name)

    nodes_data = _flatten_yaml_nodes(req_nodes)

    total, assessable = await _insert_nodes(s, fw, nodes_data)
    fw.total_nodes = total
    fw.total_assessable = assessable

    log.info(
        "Imported framework %s: %d nodes (%d assessable), URN=%s",
        fw.name, total, assessable, fw_urn,
    )

    # Create dimensions from scores_definition or score_definition
    scores_def = fw_data.get("scores_definition")
    if scores_def and isinstance(scores_def, list):
        await _create_dimensions_from_scores_definition(s, fw, scores_def)
    else:
        score_def = fw_data.get("scores") or fw_data.get("score_definition")
        if score_def and isinstance(score_def, list):
            await _create_dimensions_from_score_def(s, fw, score_def)
        else:
            await _create_default_dimensions(s, fw)

    # Create initial version record
    s.add(FrameworkVersionHistory(
        framework_id=fw.id,
        edit_version=1,
        lifecycle_status="published",
        change_summary=f"Import z YAML: {fw.name}",
        changed_by=imported_by,
        snapshot_nodes_count=total,
        snapshot_assessable_count=assessable,
    ))

    return fw


def _flatten_yaml_nodes(nodes: list[dict], depth: int = 1) -> list[dict]:
    """Flatten YAML nodes into a flat list with depth and parent_urn info.

    Handles both:
      - Nested format (children key)
      - Flat format with parent_urn and explicit depth (CISO Assistant current)
    """
    result = []
    for node in nodes:
        flat = {
            "urn": node.get("urn"),
            "parent_urn": node.get("parent_urn"),
            "ref_id": node.get("ref_id"),
            "name": node.get("name", ""),
            "description": node.get("description"),
            "depth": node.get("depth", depth),
            "assessable": node.get("assessable", False),
            "implementation_groups": node.get("implementation_groups"),
            "annotation": node.get("annotation"),
            "typical_evidence": node.get("typical_evidence"),
            "threats": node.get("threats"),
            "reference_controls": node.get("reference_controls"),
        }
        result.append(flat)

        # Support nested children format
        children = node.get("children", [])
        if children:
            # Set parent_urn for children if they don't have one
            parent_urn = node.get("urn")
            for child in children:
                if not child.get("parent_urn") and parent_urn:
                    child["parent_urn"] = parent_urn
            result.extend(_flatten_yaml_nodes(children, depth + 1))

    return result


# ===================================================
# SHARED: Insert nodes from flat list
# ===================================================

async def _insert_nodes(
    s: AsyncSession, fw: Framework, nodes_data: list[dict]
) -> tuple[int, int]:
    """Insert framework nodes. Returns (total, assessable) counts.

    Uses TWO-PASS approach for robust parent resolution:
      Pass 1: Insert all nodes (without parent_id) and build URN->ID map
      Pass 2: Update parent_id using parent_urn or depth-based fallback

    This handles CISO Assistant YAML files regardless of node ordering.
    """
    if not nodes_data:
        return (0, 0)

    total = 0
    assessable = 0

    # Determine if we have parent_urn references
    has_parent_urns = any(nd.get("parent_urn") for nd in nodes_data)

    # Track nodes for parent resolution
    urn_to_id: dict[str, int] = {}  # normalized_urn -> node_id
    inserted_nodes: list[tuple[int, dict]] = []  # (node_id, original_data)
    depth_stack: dict[int, int] = {}  # depth -> most_recent_node_id_at_that_depth

    for idx, nd in enumerate(nodes_data):
        depth = int(nd.get("depth", 1))
        is_assessable = bool(nd.get("assessable", False))

        ig = nd.get("implementation_groups")
        if ig and isinstance(ig, list):
            ig = ",".join(str(x) for x in ig)
        elif ig and not isinstance(ig, str):
            ig = str(ig)

        threats = nd.get("threats")
        if threats and isinstance(threats, str):
            threats = None  # Only store if dict/list
        if isinstance(threats, list):
            threats = threats  # Keep as list (stored as JSON)

        ref_controls = nd.get("reference_controls")
        if ref_controls and isinstance(ref_controls, str):
            ref_controls = None
        if isinstance(ref_controls, list):
            ref_controls = ref_controls

        # Ensure name is not empty -- use ref_id or description as fallback
        name = (
            str(nd.get("name") or "").strip()
            or str(nd.get("ref_id") or "").strip()
            or str(nd.get("description") or "")[:200].strip()
            or f"Node {idx + 1}"
        )

        # PASS 1: Insert node without parent (we'll set it in pass 2)
        node = FrameworkNode(
            framework_id=fw.id,
            parent_id=None,  # Will be resolved in pass 2
            urn=_normalize_urn(nd.get("urn")),
            ref_id=str(nd["ref_id"]) if nd.get("ref_id") else None,
            name=name,
            name_pl=nd.get("name_pl"),
            description=nd.get("description"),
            description_pl=nd.get("description_pl"),
            depth=depth,
            order_id=idx + 1,
            assessable=is_assessable,
            implementation_groups=ig,
            annotation=nd.get("annotation"),
            threats=threats if isinstance(threats, (dict, list)) else None,
            reference_controls=ref_controls if isinstance(ref_controls, (dict, list)) else None,
            typical_evidence=nd.get("typical_evidence"),
        )
        s.add(node)
        await s.flush()  # Get the ID

        # Build lookup maps
        node_urn = _normalize_urn(nd.get("urn"))
        if node_urn:
            urn_to_id[node_urn] = node.id

        inserted_nodes.append((node.id, nd))
        depth_stack[depth] = node.id

        total += 1
        if is_assessable:
            assessable += 1

    # PASS 2: Resolve parent_id for all nodes
    depth_stack_pass2: dict[int, int] = {}

    for node_id, nd in inserted_nodes:
        depth = int(nd.get("depth", 1))
        parent_id = None

        if has_parent_urns:
            parent_urn = _normalize_urn(nd.get("parent_urn"))
            if parent_urn:
                parent_id = urn_to_id.get(parent_urn)
                if parent_id is None:
                    log.warning(
                        "Parent URN '%s' not found for node '%s' (urn=%s)",
                        parent_urn, nd.get("ref_id") or nd.get("name"), nd.get("urn"),
                    )
            # Fallback to depth-based if parent_urn lookup failed
            if parent_id is None and depth > 1:
                parent_id = depth_stack_pass2.get(depth - 1)
        else:
            # Pure depth-based parenting
            if depth > 1:
                parent_id = depth_stack_pass2.get(depth - 1)

        if parent_id is not None:
            node_obj = await s.get(FrameworkNode, node_id)
            if node_obj:
                node_obj.parent_id = parent_id

        depth_stack_pass2[depth] = node_id

    await s.flush()

    return (total, assessable)
